
import platform
import datetime
from os.path import basename

import os
import sys
import openpyxl
from pathlib import Path
import ftplib
from ftplib import FTP
import logging

sys.path.append(os.path.join(os.getcwd(), "python"))
sys.path.append(os.path.join(os.getcwd(), "shared"))

#logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(message)s')
logging.basicConfig(format='%(asctime)s - %(levelname)s: %(message)s', 
                    level=getattr(logging, os.getenv('LOGGING_LEVEL').upper()))

def main():
  the_path_in = os.getenv('CITMAN_DATA') + '/template/in'
  the_path_out = os.getenv('CITMAN_DATA') + '/template/out'

  my_sftp_Hostname = os.getenv('SFTP_HOSTNAME') 
  my_sftp_Username = os.getenv('SFTP_USERNAME') 
  my_sftp_Password = os.getenv('SFTP_PASSWORD') 
  my_sftp_path_main = os.getenv('SFTP_PATH') 
  
  # get the fril from /groupwide_shares/Group_Restricted_Data/Collections/Shared
  my_sftp_path_in = my_sftp_path_main + '/Template_Import/template_in'
  logging.debug('sftp_path_in = ' + my_sftp_path_in)
  logging.debug('the_path_in = ' + the_path_in)
  
  current_dir = os.getcwd()
  logging.debug("current_dir: " + str(current_dir))
  logging.debug(" ")
  try:

        with FTP(my_sftp_Hostname) as ftp:
            ftp.login(user=my_sftp_Username, passwd=my_sftp_Password)
            logging.debug("FTP succesfully stablished ... ")
            ftp.cwd(my_sftp_path_in)
            logging.debug("FTP cwd succesfully stablished ... ")
            #os.chdir(".\data\template\in")
            os.chdir(the_path_in)
            filelist = ftp.nlst("*.xlsx")
            for file in filelist:
                logging.debug("File to upload: " + str(file))
                localfile = open(file, 'wb')
                ftp.retrbinary('RETR ' + file, localfile.write, 1024)
                localfile.close()
            logging.debug("call ftp.quit")
            ftp.quit()
  except:
        logging.error("FTP connection failure")
  logging.debug(" ")
  
  os.chdir(current_dir)
  the_files = os.listdir(the_path_in)
  for the_file in the_files:
    logging.debug("in data: " + str(the_file))
    xlsx_file = Path(the_path_in, the_file)
    logging.debug("Handling of the file: " + str(xlsx_file))

    from openpyxl import load_workbook
    wb = openpyxl.load_workbook(xlsx_file)
    wsTemplate = wb["BU - Template"]
    wsFees = wb["Template - Inv. Items - Fees"]

    boOK = True
    error_dict = {"Errors":[]}

    iBU = 2
    
    while str(wsTemplate.cell(row=iBU, column=1).value).strip() != 'None' and boOK == True:
       
       l_BU = str(wsTemplate.cell(row=iBU, column=1).value).strip()
       l_template = str(wsTemplate.cell(row=iBU, column=2).value).strip()
       
       if len(l_template) > 20:
          boOK = False
          error_dict["Errors"].append({'message': 'The size of the template cannot exceed 20 positions (' + l_template + ')'})

       l_currency = str(wsTemplate.cell(row=iBU, column=3).value).strip()
       
       logging.debug('l_BU: ' + str(l_BU) + ' - template: ' + l_template + ' - currency: ' + l_currency)
       
       # handling of the fee_codes
       iFC = 2
       unique = []
  
       while str(wsFees.cell(row=iFC, column=8).value).strip() != 'None' :
          l_feeCode = str(wsFees.cell(row=iFC, column=8).value).strip()

          if l_feeCode not in unique:
             unique.append(l_feeCode)

          iFC = iFC + 1

       for l_feeCode in unique:
         logging.debug('Fee Code to ask a check: ' + l_feeCode + ' - ' + l_BU + ' - '+ l_currency)    


       # index of templates in the BU Template sheet
       iBU = iBU + 1
    
       

    wb.close()
    
    the_file_without_ext = Path(xlsx_file).stem
    logging.debug(" without ext: " + str( the_file_without_ext))
    fileNameFTP = the_file_without_ext + "_" + datetime.datetime.today().strftime('%Y%m%d-%H%M') + '.xlsx'
    logging.debug("FTP connection")
    try:
       with FTP(my_sftp_Hostname) as ftp:
         ftp.login(user=my_sftp_Username, passwd=my_sftp_Password)
         logging.debug("FTP succesfully stablished ... ")
         ftp.cwd(my_sftp_path_in)
         logging.debug("FTP cwd succesfully stablished ... ")
         ftp.delete(the_file)
         ftp.cwd('./bck')
         with open(xlsx_file, 'rb') as file:
            result = ftp.storbinary(f'STOR ' + fileNameFTP, file)
            logging.debug("result of store binary: " + str(result))
            # file.close()
            logging.debug("call ftp.quit")
            ftp.quit()
    except:
      logging.error("FTP connection failure")

    # BEVRIG1: PUT IN COMMENT TO AVOID TO RESET THE INPUT FILE EACH RUN IN J DRIVE
    # if os.path.exists(xlsx_file):
    #  os.remove(xlsx_file)
    print(" ")
    
    if boOK == False :
          logging.error(str(error_dict))
          
    logging.debug(" ")
    

if __name__ == "__main__":
    main()