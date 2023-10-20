import platform
import datetime
from os.path import basename

import cx_Oracle
import os
import configparser
import requests
import json
import sys
import openpyxl
from pathlib import Path
import ftplib
from ftplib import FTP
import logging

# setting shared path
sys.path.append(os.path.join(os.getcwd(), "shared"))

#from cit_shared.cit_env import CIT_env
#from ..shared.cit_shared.cit_set_env import CIT_set_env

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from subprocess import Popen, PIPE

cit_env = CIT_set_env()

env_params = cit_env.get_env_params()
env_data = cit_env.get_env_data()

# Read ini file and get variables
print('env_params= ' + env_params)
parser = configparser.ConfigParser()
parser.read(env_params + 'score_api.ini')
score_api_auth = parser['score_api']['score_api_auth']
score_url= parser['score_api']['score_url_add_info']

parser = configparser.ConfigParser()
parser.read(env_params + 'sftp_restricted.ini')
my_sftp_Hostname = parser['sftp_restricted']['sftp_Hostname']
my_sftp_Username = parser['sftp_restricted']['sftp_Username']
my_sftp_Password = parser['sftp_restricted']['sftp_Password']
my_destination = parser['sftp_restricted']['retrievepath_load_info']

parser = configparser.ConfigParser()
parser.read(env_params + 'auto_load_info.ini')
the_ftp_filepath = parser['scorepathmonitor']['savepath']
mail_title = parser['scorepathmonitor']['mail_title']
mail_to = parser['scorepathmonitor']['mail_to']
mail_cc = parser['scorepathmonitor']['mail_cc']
mail_from = parser['scorepathmonitor']['mail_from']

def main():

    '''
    # Count the arguments
    # in comment because not needed
    arguments = len(sys.argv) - 1
    print("Number of arguments: ", arguments)
    # Output argument-wise
    inputPathFile = 'data'
    position = 1
    while (arguments >= position):
        print("Parameter %i: %s" % (position, sys.argv[position]))
        if position == 1:
            fileName = sys.argv[position]
        elif position == 2:
            inputPathFile = sys.argv[position]
        position = position + 1
    '''
    
    print("DESCRIPTION")
    print("With SFTP, upload xls files from J drive")
    print("Call the SCORE API add_info for each line of the excel files")
    print("Delete the files from J drive and save them to a bck directory in J")
    print("Send a HTML mail with the errors detected by the SCORE API")
    print(" ")
    print("env_data: ", str(env_data))
    print("env_params: ", str(env_params))

    print("score_api_auth: ", score_api_auth)
    print("score_url", score_url)
    print("my_destination: ", my_destination)
    print("the_ftp_filepath: ", the_ftp_filepath)
    print("my_sftp_Hostname: ", my_sftp_Hostname)
    print("my_sftp_Username: ", my_sftp_Username)
    print("my_sftp_Password: ", my_sftp_Password)
    print("mail_title: ", mail_title)
    print("mail_to: ", mail_to)
    print("mail_cc: ", mail_cc)
    print("mail_from: ", mail_from)




    current_dir = os.getcwd()
    print("current_dir: ", current_dir)
    print(" ")

    try:

        with FTP(my_sftp_Hostname) as ftp:
            ftp.login(user=my_sftp_Username, passwd=my_sftp_Password)
            print("FTP succesfully stablished ... ")
            ftp.cwd(my_destination)
            print("FTP cwd succesfully stablished ... ")
            #os.chdir("./data/info_from_ftp")
            os.chdir(the_ftp_filepath)
            filelist = ftp.nlst("*.xlsx")
            for file in filelist:
                print("File to upload: ", file)
                localfile = open(file, 'wb')
                ftp.retrbinary('RETR ' + file, localfile.write, 1024)
                localfile.close()
            print("call ftp.quit")
            ftp.quit()
    except:
        print("FTP connection failure")
    print(" ")

    os.chdir(current_dir)
    the_files = os.listdir(the_ftp_filepath)
    for the_file in the_files:
        print("in data: ", the_file)
        xlsx_file = Path(the_ftp_filepath, the_file)
        print("Handling of the file: ", xlsx_file)

        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        print(sheet)

        max_row = sheet.max_row
        print("maxRow: ", max_row)
        bo_error = False
        error_dict = {"Errors": []}
        
        #max_col = sheet.max_column
        #print("max_col", max_col)
        if max_row > 1:
          # get token to use
          print("Get token")
          response = requests.get(score_api_auth)
          if response.status_code == 200 or response.status_code == 201:
            body = response.json()
            #print("body: ", body)
            sessionKey = body['sessionKey']
            print("sessionKey: ", sessionKey)
            v_url_to_call = score_url

            print("Read xlsx file")
            for i in range(2, max_row + 1):
                #refdoss = str(sheet.cell_value(i, 0)).replace('.0', '')
                print(i)
                cell_obj = sheet.cell(row=i, column=1)
                refdoss = str(cell_obj.value).strip()

                cell_obj = sheet.cell(row=i, column=2)
                comment = cell_obj.value
                cell_obj = sheet.cell(row=i, column=3)
                sta = str(cell_obj.value).strip()
                if (
                     comment is not None and comment != '' and comment != 'None'
                   ) \
                   or \
                   (
                    sta is not None and sta != '' and sta != 'None'
                   ):
                  if refdoss is not None and refdoss != '':
                    print("refdoss: ", refdoss)
                  if comment is not None and comment != '' and comment != 'None':
                    print("comment: ", comment.encode('utf-8'))
                  if sta is not None and sta != '' and sta != 'None':
                    print("sta: ", sta)

                  # handling of the STA
                  if sta is not None and sta != '' and sta != 'None':

                    today = str(datetime.date.today())
                    infoload = {"casRef": refdoss,
                                "inDat": today,
                                "inMsgType": sta
                               }

                    infoloadAPI = json.dumps(infoload)

                    print("Call API info STA: ", str(infoloadAPI))

                    try:
                      headers = {'Content-Type': 'application/json',
                                 'iMX-Session-Key': sessionKey}

                      print("Header: ", str(headers))
                      response = requests.post(v_url_to_call, headers=headers, data=infoloadAPI.encode('utf-8'))
                      body = response.json()

                      print("Response: ", str(response))
                      print("Response Body: ", str(body))

                      if response.status_code == 200 or response.status_code == 201:
                          print("in sta inserted - ref:", str(body["inRef"]))
                      else :
                          print("ERROR to save this STA...")
                          bo_error = True
                          #print("error:", str(body['message']))
                          error_message = {'FileName': the_file, 'Case': str(refdoss)
                                          , 'Type': 'STA', 'error': body['message']
                                          ,'details' : body['detail']}
                          print("error_message: ", str(error_message))
                          error_dict["Errors"].append(error_message)
                    except:
                        error_message = "Cannot call the API to save in STA"

                  # handling of the comment
                  if comment is not None and comment != '' and comment != 'None':

                    # max length of the free_text is 1000
                    max_length = 1000
                    comment = str(comment)
                    today = str(datetime.date.today())
                    while comment != '' and comment is not None:
                      #print ("comment: ", comment)
                      print ("comment: ", comment.encode('utf-8'))
                      comment_max = comment[0: max_length]
                      comment = comment[max_length:]
                      infoload = {"casRef": refdoss,
                                  "inDat": today,
                                  "inFreeText": comment_max
                                 }

                      infoloadAPI = json.dumps(infoload)

                      print("Call API info STA: ", str(infoloadAPI))

                      try:
                        headers = {'Content-Type': 'application/json',
                                   'iMX-Session-Key': sessionKey}

                        print("Header: ", str(headers))
                        response = requests.post(v_url_to_call, headers=headers, data=infoloadAPI.encode('utf-8'))
                        body = response.json()

                        print("Response: ", str(response))
                        print("Response Body: ", str(body))

                        if response.status_code == 200 or response.status_code == 201:
                          print("in free text inserted - ref:", str(body["inRef"]))
                        else:
                          print("ERROR to save this Free Text...")
                          bo_error = True
                          error_message = {'FileName': the_file, 'Case': str(refdoss), 'Type': 'FreeText'
                                          , 'error': body['message']
                                          ,'details' : body['detail']}
                          print("error_message: ", str(error_message))
                          error_dict["Errors"].append(error_message)
                      except:
                        error_message = "Cannot call the API to save in free text"

                else :
                  print("No entry for this line")

        wb_obj.close()
        print(" ")

        the_file_without_ext = Path(xlsx_file).stem
        print(" without ext: ", the_file_without_ext)
        fileNameFTP = the_file_without_ext + "_" + datetime.datetime.today().strftime('%Y%m%d-%H%M') + '.xlsx'
        print("FTP connection")
        try:
            with FTP(my_sftp_Hostname) as ftp:
                ftp.login(user=my_sftp_Username, passwd=my_sftp_Password)
                logging.info("FTP succesfully stablished ... ")
                ftp.cwd(my_destination)
                logging.info("FTP cwd succesfully stablished ... ")
                ftp.delete(the_file)
                ftp.cwd('./bck')
                with open(xlsx_file, 'rb') as file:
                  result = ftp.storbinary(f'STOR ' + fileNameFTP, file)
                  print("result of store binary: ", str(result))
                  # file.close()
                logging.info("call ftp.quit")
                ftp.quit()
        except:
            logging.info("FTP connection failure")

        if os.path.exists(xlsx_file):
            os.remove(xlsx_file)
        print(" ")

        if bo_error == True:

          the_error = json.dumps(error_dict)
          print("Error returned: " + str(the_error))

          msg = MIMEMultipart("alternative")
          msg["From"] = mail_from
          msg["Subject"] = mail_title + " - File: " + the_file
          msg["To"] = mail_to
          msg["Cc"] = mail_cc

          html = """<html>
                  <head>
                      <meta charset=utf-8" />
                      <title>Information auto load error</title>
                      <style>
                      body {font: 12px Arial}
                      table {margin-left: 30px; border-collapse: collapse; font: Calibri; background-colour: #c0c0c0}
                      caption {background-color: white; border: 0; caption-side: bottom; font: bold 10px}
                      th {background-color: #c0c0c0; color: white; padding: 6px 4px; text-align: center}
                     td {border: 1px solid #95b3d7;vertical-align: top; padding: 4px }
                      </style>
                  </head>
                  <body>
                      <p>Hello, </p>
                      <p>Please find below list of information rejected during the process of file """
          html += the_file + """:</p>"""
          html += """
                      <b>Rejected movements</b><p></p>
                      <table>
                          <tr><th>File</th><th>Case</th><th>Type</th><th>Error</th><th>Details</th>"""
          for _calls in error_dict["Errors"]:
            html += '<tr><td>' + _calls["FileName"] + '</td><td>' + _calls["Case"] + '</td><td>' + _calls["Type"] + \
                          '</td><td>' + _calls["error"] + '</td><td>' + _calls["details"] + '</td></tr>'
          html += '</table>'
          html += """<br><br><p>This mail is sent automatically.</p>
                      <p>Please do not answer to this mail. However do not hesitate to contact <a href="mailto:support.collections@atradius.com">Collections Support</a> for any question.</p>
                      <p>Regards<br><br><img src="https://www.atradius.com/ATRADIUS/help/gif/Atradius_Collections.gif" /></p>
                  </body></html>"""

          msg.attach(MIMEText(html, "html"))

          if platform.system() == 'Windows':
              print("Cannot send email with Python from Windows, error details:")
              for _calls in error_dict["Errors"]:
                  print('error FileName: ', _calls["FileName"])
                  print('- case: ', _calls["Case"])
                  print('- type: ', _calls["Type"])
                  print('- error: ', _calls["error"])
                  print('- error details: ', _calls["details"])
          else:
            p = Popen(["/usr/sbin/sendmail", "-t"], stdin=PIPE)
            p.communicate(msg.as_string().encode())
            print('Mail sent to', msg['To'])
    print(" ")

    # end of loop of the files

if __name__ == "__main__":
    main()

